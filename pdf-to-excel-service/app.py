from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill
import io
import os

app = Flask(__name__)

# Configurar CORS para permitir comunicación desde Vercel
CORS(app, origins=[
    "https://acubat-pricing-platform.vercel.app",  # Tu dominio Vercel real
    "https://acubat-pricing-platform-git-main-ralborta.vercel.app",  # Dominio de preview
    "http://localhost:3000",  # Para desarrollo local
    "https://localhost:3000"  # Para desarrollo local con HTTPS
])

@app.route('/health', methods=['GET'])
def health_check():
    """Endpoint de salud para Railway"""
    return jsonify({
        "status": "healthy",
        "service": "pdf-to-excel",
        "version": "1.0.0"
    })

@app.route('/convert', methods=['POST'])
def convert_pdf_to_excel():
    """Convertir PDF a Excel"""
    try:
        # Verificar si hay archivo en la request
        if 'file' not in request.files:
            return jsonify({"error": "No se proporcionó archivo"}), 400
        
        file = request.files['file']
        
        # Verificar si el archivo es un PDF
        if file.filename == '' or not file.filename.lower().endswith('.pdf'):
            return jsonify({"error": "Archivo debe ser un PDF"}), 400
        
        # Leer el PDF
        pdf_content = file.read()
        
        # Extraer texto del PDF
        text_content = []
        with pdfplumber.open(io.BytesIO(pdf_content)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    text_content.append(text)
        
        # Crear archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PDF_Content"
        
        # Estilo para el header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Escribir contenido en Excel
        row = 1
        for page_num, page_text in enumerate(text_content, 1):
            # Header de página
            ws[f'A{row}'] = f"Página {page_num}"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            # Contenido de la página
            lines = page_text.split('\n')
            for line in lines:
                if line.strip():
                    ws[f'A{row}'] = line.strip()
                    row += 1
            
            # Espacio entre páginas
            row += 1
        
        # Ajustar ancho de columna
        ws.column_dimensions['A'].width = 100
        
        # Guardar Excel en memoria
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=f"{os.path.splitext(file.filename)[0]}_converted.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        return jsonify({"error": f"Error en la conversión: {str(e)}"}), 500

@app.route('/', methods=['GET'])
def home():
    """Página de inicio del microservicio"""
    return jsonify({
        "service": "PDF to Excel Converter",
        "version": "1.0.0",
        "endpoints": {
            "health": "/health",
            "convert": "/convert (POST)"
        }
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
