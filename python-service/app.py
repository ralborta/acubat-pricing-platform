from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill
import io
import tempfile
import os
from datetime import datetime

app = Flask(__name__)

# Configurar CORS para permitir solo tu dominio Vercel
CORS(app, origins=[
    "https://tu-proyecto.vercel.app",  # Cambiar por tu dominio real
    "http://localhost:3000"            # Para desarrollo local
])

@app.route('/convert-pdf', methods=['POST'])
def convert_pdf():
    try:
        # Verificar que se envió un archivo
        if 'file' not in request.files:
            return jsonify({'error': 'No se proporcionó archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        # Verificar que sea un PDF
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'El archivo debe ser un PDF'}), 400
        
        print(f"Procesando PDF: {file.filename}")
        
        # Extraer tablas del PDF con pdfplumber
        tables = []
        with pdfplumber.open(file) as pdf:
            for page_num, page in enumerate(pdf.pages):
                print(f"Procesando página {page_num + 1}")
                
                # Extraer tablas de la página
                page_tables = page.extract_tables()
                if page_tables:
                    tables.extend(page_tables)
                    print(f"Encontradas {len(page_tables)} tablas en página {page_num + 1}")
                
                # Si no hay tablas, extraer texto como fallback
                if not page_tables:
                    text = page.extract_text()
                    if text:
                        # Crear tabla simple con el texto
                        lines = text.strip().split('\n')
                        table_data = [[line] for line in lines if line.strip()]
                        if table_data:
                            tables.append(table_data)
                            print(f"Texto extraído como tabla en página {page_num + 1}")
        
        if not tables:
            return jsonify({'error': 'No se encontraron tablas ni texto en el PDF'}), 400
        
        print(f"Total de tablas encontradas: {len(tables)}")
        
        # Crear Excel con openpyxl
        wb = Workbook()
        
        # Crear hoja para cada tabla
        for i, table in enumerate(tables):
            if i == 0:
                ws = wb.active
                ws.title = f"Tabla_{i+1}"
            else:
                ws = wb.create_sheet(f"Tabla_{i+1}")
            
            # Insertar datos de la tabla
            for row_idx, row in enumerate(table, 1):
                for col_idx, cell_value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    
                    # Aplicar estilos
                    if row_idx == 1:  # Encabezados
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:  # Datos
                        cell.alignment = Alignment(vertical='center')
                    
                    # Aplicar bordes
                    cell.border = Border(
                        left=Border(style='thin'),
                        right=Border(style='thin'),
                        top=Border(style='thin'),
                        bottom=Border(style='thin')
                    )
            
            # Ajustar ancho de columnas automáticamente
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Generar Excel en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre del archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"converted_pdf_{timestamp}.xlsx"
        
        print(f"Excel generado exitosamente: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"Error en conversión: {str(e)}")
        return jsonify({'error': f'Error en la conversión: {str(e)}'}), 500

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'healthy',
        'service': 'pdf-to-excel',
        'timestamp': datetime.now().isoformat()
    })

@app.route('/', methods=['GET'])
def root():
    return jsonify({
        'message': 'PDF to Excel Microservice',
        'endpoints': {
            'convert': '/convert-pdf',
            'health': '/health'
        }
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    print(f"Iniciando microservicio en puerto {port}")
    app.run(host='0.0.0.0', port=port, debug=False)
